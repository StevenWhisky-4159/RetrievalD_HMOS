#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""将目录映射记录写出为 Excel。"""
from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from .parser import KitRecord
except ImportError:
    from parser import KitRecord

KIT_COLUMNS = ("领域", "中文", "英文", "对应指南", "basic skill", "API参考")
KIT_SHEET_NAME = "Kit路由"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Microsoft YaHei", size=10)
ALT_FILL = PatternFill("solid", fgColor="D6E3F0")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
WRAP = Alignment(vertical="center", wrap_text=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

MIN_WIDTHS = {
    "领域": 16,
    "所属主题": 28,
    "层级": 8,
    "类型": 10,
    "中文": 22,
    "英文": 28,
    "对应指南": 56,
    "basic skill": 56,
    "API参考": 52,
    "Markdown标题": 48,
    "kit路由": 10,
}
MAX_WIDTH = 80


def _display_width(text: str) -> float:
    width = 0.0
    for char in text:
        width += 2.1 if ord(char) > 127 else 1.05
    return width


def _column_width(header: str, values: list[str]) -> float:
    widths = [_display_width(header)] + [_display_width(value) for value in values]
    fitted = min(max(widths) + 2, MAX_WIDTH)
    return max(fitted, MIN_WIDTHS.get(header, 12))


def write_table_excel(
    rows: list[dict[str, str]],
    output_path: Path,
    *,
    columns: Sequence[str],
    sheet_name: str,
    wrap_body: bool | None = None,
) -> Path:
    if not rows:
        raise ValueError("没有可写出的记录")

    wrap_body = wrap_body if wrap_body is not None else len(rows) <= 400
    wrap = WRAP if wrap_body else Alignment(vertical="center")
    style_body = len(rows) <= 400

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(list(columns))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN

    for index, row in enumerate(rows, start=2):
        ws.append([row.get(column, "") for column in columns])
        if not style_body:
            continue
        fill = ALT_FILL if index % 2 == 0 else None
        for cell in ws[index]:
            cell.font = BODY_FONT
            cell.alignment = wrap
            cell.border = THIN
            if fill is not None:
                cell.fill = fill

    if "Markdown标题" in columns:
        title_column = columns.index("Markdown标题") + 1
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=title_column).alignment = WRAP

    for col_idx, header in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        if style_body:
            values = [str(ws.cell(row=row_idx, column=col_idx).value or "") for row_idx in range(2, ws.max_row + 1)]
            ws.column_dimensions[letter].width = _column_width(header, values)
        else:
            ws.column_dimensions[letter].width = MIN_WIDTHS.get(header, 18)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    ws.sheet_properties.tabColor = "1F4E79"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def write_kit_excel(records: list[KitRecord], output_path: Path) -> Path:
    return write_table_excel(
        [record.to_row() for record in records],
        output_path,
        columns=KIT_COLUMNS,
        sheet_name=KIT_SHEET_NAME,
    )
